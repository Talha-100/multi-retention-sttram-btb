import openpyxl

wb = openpyxl.load_workbook('Analysis_Motivation.xlsx')
for sheetname in wb.sheetnames:
    print(f"Sheet: {sheetname}")
    sheet = wb[sheetname]
    for row in range(1, 50):
        row_vals = []
        for col in range(1, 30):
            val = sheet.cell(row=row, column=col).value
            row_vals.append(str(val))
        if any(v != "None" for v in row_vals):
            print(f"Row {row}: " + "\t".join(row_vals))
