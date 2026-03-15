import pandas as pd
import sys
import os
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image

def plot_spatial_distribution(df_plot, img_path):
    if df_plot is None or df_plot.empty: return
    df_plot = df_plot.copy()
    if 'Set' in df_plot.columns:
        df_plot['Set'] = df_plot['Set'].astype(int)
        w_cols = [c for c in df_plot.columns if str(c).startswith('W')]
        if w_cols:
            df_plot = df_plot[df_plot[w_cols].sum(axis=1) > 0]
            if len(df_plot) > 10:
                indices = np.linspace(0, len(df_plot) - 1, 10, dtype=int)
                df_plot = df_plot.iloc[indices]
            df_plot = df_plot.set_index('Set')
            
            fig, ax = plt.subplots(figsize=(10, 5))
            colors = ['#003f7f', '#ff4c00', '#ffcc00', '#4c9900', 'white', '#7fccff', 'white', '#99cc00', '#ffcc99']
            hatches = ['', '', '', '', '---', '', '/', '', '+++']
            n_cols = len(w_cols)
            colors = (colors * ((n_cols // len(colors)) + 1))[:n_cols]
            hatches = (hatches * ((n_cols // len(hatches)) + 1))[:n_cols]

            x = np.arange(len(df_plot.index))
            width = 0.8 / n_cols
            
            for i, col in enumerate(w_cols):
                offset = (i - n_cols/2) * width + width/2
                edgecolor = 'black' if hatches[i] != '' or colors[i] == 'white' else 'none'
                ax.bar(x + offset, df_plot[col], width, label=col, color=colors[i], edgecolor=edgecolor, hatch=hatches[i], linewidth=1)

            ax.set_ylabel('Write Count', fontweight='bold', fontsize=12)
            ax.set_xlabel('Set ID', fontweight='bold', fontsize=12)
            ax.set_xticks(x)
            ax.set_xticklabels(df_plot.index, fontweight='bold', fontsize=11)
            
            import matplotlib.patches as mpatches
            legend_handles = []
            for i, col in enumerate(w_cols):
                edgecolor = 'black' if hatches[i] != '' or colors[i] == 'white' else 'none'
                patch = mpatches.Patch(facecolor=colors[i], edgecolor=edgecolor, hatch=hatches[i], label=col, linewidth=1)
                legend_handles.append(patch)

            ax.legend(handles=legend_handles, loc='upper center', bbox_to_anchor=(0.5, 1.15), 
                      ncol=min(n_cols, 9), frameon=False, prop={'weight':'bold', 'size':11}, handletextpad=0.5, columnspacing=1.0)

            ax.yaxis.grid(True, linestyle='-', which='major', color='lightgrey', alpha=0.5)
            ax.set_axisbelow(True)
            ax.spines['top'].set_visible(False)
            ax.spines['right'].set_visible(False)
            ax.spines['left'].set_linewidth(1.5)
            ax.spines['bottom'].set_linewidth(1.5)
            ax.tick_params(axis='y', labelsize=11, width=1.5, length=5)
            ax.tick_params(axis='x', length=5, width=1.5)
            
            plt.tight_layout()
            plt.savefig(img_path, dpi=300)
            plt.close()

def analyze_stt_writes(input_file, excel_file):
    print(f"Analyzing STT-RAM writes from {input_file}...")
    
    data = []
    try:
        with open(input_file, 'r') as f:
            for line in f:
                parts = line.split()
                if len(parts) >= 7 and parts[2] == 'STT_WRITE_COUNT':
                    benchmark = parts[0]
                    config = parts[1]
                    partition = int(parts[3])
                    set_idx = int(parts[4])
                    way = int(parts[5])
                    count = int(parts[6])
                    
                    data.append({
                        'Benchmark': benchmark,
                        'Config': config,
                        'Partition': partition,
                        'Set': set_idx,
                        'Way': way,
                        'Count': count
                    })
    except FileNotFoundError:
        print(f"Error: Input file '{input_file}' not found.")
        return

    if not data:
        print("No STT_WRITE_COUNT data found.")
        return

    df = pd.DataFrame(data)
    
    partition_map = {
        0: "Return", 1: "<= 4 bits", 2: "<= 5 bits", 3: "<= 7 bits",
        4: "<= 9 bits", 5: "<= 11 bits", 6: "<= 19 bits", 7: "<= 25 bits", 8: "> 25 bits"
    }

    def generate_table(df_subset):
        if df_subset.empty: return pd.DataFrame()
        stats = df_subset.groupby('Partition')['Count'].sum().reset_index()
        stats.columns = ['Partition ID', 'Total Writes']
        stats['Offset Bits'] = stats['Partition ID'].map(partition_map)
        total_sum = stats['Total Writes'].sum()
        stats['% of Total'] = (stats['Total Writes'] / total_sum * 100) if total_sum > 0 else 0.0
        return stats[['Partition ID', 'Offset Bits', 'Total Writes', '% of Total']]

    def generate_spatial_table(df_subset):
        if df_subset.empty: return pd.DataFrame()
        spatial = df_subset.groupby(['Set', 'Partition'])['Count'].sum().unstack(fill_value=0)
        spatial.columns = [f"W{c}" for c in spatial.columns]
        return spatial.reset_index()

    unique_configs = df['Config'].unique()
    no_sttram_configs = [c for c in unique_configs if 'no' in c and 'sttramBTB' in c]
    fdip_sttram_configs = [c for c in unique_configs if 'fdip' in c and 'sttramBTB' in c]
    no_fixed_configs = [c for c in unique_configs if 'no' in c and 'fixed-retentions-btb' in c]
    fdip_fixed_configs = [c for c in unique_configs if 'fdip' in c and 'fixed-retentions-btb' in c]
    
    tables = {}
    spatial_tables = {}
    
    if no_sttram_configs:
        df_sub = df[df['Config'].isin(no_sttram_configs)]
        tables['no_sttramBTB Stats'] = generate_table(df_sub)
        spatial_tables['no_sttramBTB Spatial'] = generate_spatial_table(df_sub)
    if fdip_sttram_configs:
        df_sub = df[df['Config'].isin(fdip_sttram_configs)]
        tables['fdip_sttramBTB Stats'] = generate_table(df_sub)
        spatial_tables['fdip_sttramBTB Spatial'] = generate_spatial_table(df_sub)
    if no_fixed_configs:
        df_sub = df[df['Config'].isin(no_fixed_configs)]
        tables['no_fixed-retentions-btb Stats'] = generate_table(df_sub)
        spatial_tables['no_fixed-retentions-btb Spatial'] = generate_spatial_table(df_sub)
    if fdip_fixed_configs:
        df_sub = df[df['Config'].isin(fdip_fixed_configs)]
        tables['fdip_fixed-retentions-btb Stats'] = generate_table(df_sub)
        spatial_tables['fdip_fixed-retentions-btb Spatial'] = generate_spatial_table(df_sub)

    if 'no_sttramBTB Stats' in tables and 'fdip_sttramBTB Stats' in tables:
        t1, t2 = tables['no_sttramBTB Stats'].set_index('Partition ID'), tables['fdip_sttramBTB Stats'].set_index('Partition ID')
        avg_df = t1[['Total Writes']].add(t2[['Total Writes']], fill_value=0) / 2
        avg_df = avg_df.reset_index()
        avg_df.columns = ['Partition ID', 'Average Total Writes']
        avg_df['Offset Bits'] = avg_df['Partition ID'].map(partition_map)
        tot = avg_df['Average Total Writes'].sum()
        avg_df['% of Total'] = (avg_df['Average Total Writes'] / tot * 100) if tot > 0 else 0.0
        tables['sttramBTB Combined Average Stats'] = avg_df[['Partition ID', 'Offset Bits', 'Average Total Writes', '% of Total']]
        
        s1, s2 = spatial_tables['no_sttramBTB Spatial'].set_index('Set'), spatial_tables['fdip_sttramBTB Spatial'].set_index('Set')
        spatial_tables['sttramBTB Combined Average Spatial'] = s1.add(s2, fill_value=0) / 2

    if 'no_fixed-retentions-btb Stats' in tables and 'fdip_fixed-retentions-btb Stats' in tables:
        t1, t2 = tables['no_fixed-retentions-btb Stats'].set_index('Partition ID'), tables['fdip_fixed-retentions-btb Stats'].set_index('Partition ID')
        avg_df = t1[['Total Writes']].add(t2[['Total Writes']], fill_value=0) / 2
        avg_df = avg_df.reset_index()
        avg_df.columns = ['Partition ID', 'Average Total Writes']
        avg_df['Offset Bits'] = avg_df['Partition ID'].map(partition_map)
        tot = avg_df['Average Total Writes'].sum()
        avg_df['% of Total'] = (avg_df['Average Total Writes'] / tot * 100) if tot > 0 else 0.0
        tables['fixed-retentions-btb Combined Average Stats'] = avg_df[['Partition ID', 'Offset Bits', 'Average Total Writes', '% of Total']]
        
        s1, s2 = spatial_tables['no_fixed-retentions-btb Spatial'].set_index('Set'), spatial_tables['fdip_fixed-retentions-btb Spatial'].set_index('Set')
        spatial_tables['fixed-retentions-btb Combined Average Spatial'] = s1.add(s2, fill_value=0).reset_index()

    print(f"Writing to {excel_file}...")
    try:
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
             start_row = 0
             for title, table_df in tables.items():
                 pd.DataFrame([title]).to_excel(writer, sheet_name='sttramBTB writes', startrow=start_row, index=False, header=False)
                 start_row += 1
                 table_df.to_excel(writer, sheet_name='sttramBTB writes', startrow=start_row, index=False)
                 start_row += len(table_df) + 3
                 
             start_row = 0
             for title, table_df in spatial_tables.items():
                 pd.DataFrame([title]).to_excel(writer, sheet_name='Spatial Distribution', startrow=start_row, index=False, header=False)
                 start_row += 1
                 table_df.to_excel(writer, sheet_name='Spatial Distribution', startrow=start_row, index=False)
                 start_row += len(table_df) + 3

        print(f"Successfully created {excel_file} with analysis.")
    except Exception as e:
        print(f"Error writing Excel: {e}")
        return

    # Generate and embed the spatial plot
    plot_df = None
    if 'fixed-retentions-btb Combined Average Spatial' in spatial_tables:
         plot_df = spatial_tables['fixed-retentions-btb Combined Average Spatial']

    if plot_df is not None and not plot_df.empty:
         print("Generating spatial plot...")
         img_path = 'spatial_distribution.png'
         plot_spatial_distribution(plot_df, img_path)
         
         try:
             wb = load_workbook(excel_file)
             if 'Spatial Distribution' in wb.sheetnames:
                 ws = wb['Spatial Distribution']
                 img = Image(img_path)
                 # Place the image to the right of the tables (e.g. column M)
                 ws.add_image(img, 'M2')
                 wb.save(excel_file)
                 print("Plot successfully embedded into the Excel file.")
         except Exception as e:
             print(f"Error embedding plot: {e}")

    print("Generating individual benchmark spatial plots...")
    output_dir = "spatial_plots"
    os.makedirs(output_dir, exist_ok=True)
    fixed_configs = no_fixed_configs + fdip_fixed_configs
    if fixed_configs:
        df_fixed = df[df['Config'].isin(fixed_configs)]
        for bench, group in df_fixed.groupby('Benchmark'):
            spatial_table = group.groupby(['Set', 'Partition'])['Count'].mean().unstack(fill_value=0)
            spatial_table.columns = [f"W{c}" for c in spatial_table.columns]
            spatial_table = spatial_table.reset_index()
            if not spatial_table.empty:
                b_img_path = os.path.join(output_dir, f"{bench}_spatial_distribution.png")
                plot_spatial_distribution(spatial_table, b_img_path)

if __name__ == "__main__":
    input_filename = "collectStats/all_res"
    output_filename = "sttram_analysis_results.xlsx" 
    
    if len(sys.argv) > 1:
        input_filename = sys.argv[1]
    if len(sys.argv) > 2:
        output_filename = sys.argv[2]

    analyze_stt_writes(input_filename, output_filename)