import pandas as pd
import openpyxl
import os
import sys

def parse_all_res(filepath):
    data = []
    try:
        with open(filepath, 'r') as f:
            for line in f:
                if 'MOT_STATS' in line:
                    parts = line.strip().split()
                    # Expected format: bench_name runName_btb MOT_STATS trace_string zone bucket count
                    # Example: server_001 no_convBTB MOT_STATS server_001.champsimtrace.xz 0 1 500
                    if len(parts) >= 7 and parts[2] == 'MOT_STATS':
                        config = parts[1]
                        
                        # Only collect for convBTB without prefetcher (baseline SRAM BTB profiling)
                        if config == 'no_convBTB':
                            bench = parts[0]
                            try:
                                zone = int(parts[4])
                                bucket = int(parts[5])
                                count = int(parts[6])
                                data.append([bench, zone, bucket, count])
                            except ValueError:
                                pass
    except FileNotFoundError:
        print(f"Error: {filepath} not found")
        return pd.DataFrame()
        
    if not data:
        return pd.DataFrame()
        
    df = pd.DataFrame(data, columns=['Benchmark', 'Zone', 'Bucket', 'Count'])
    df = df.groupby(['Benchmark', 'Zone', 'Bucket'], as_index=False)['Count'].sum()
    print(f"Parsed {len(df)} records for {len(df['Benchmark'].unique())} unique benchmarks")
    print(f"Unique benchmarks in data: {df['Benchmark'].unique()}")
    return df

def update_excel(excel_path, df):
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        print(f"Error: {excel_path} not found")
        return
        
    print(f"Sheet names: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # Find benchmark rows
        found_benches = []
        last_row = 1
        print(f"Scanning rows in {sheet_name} for benchmarks...")
        for row in range(1, 2000):
            # The benchmark name might be in column 14 (N) or 15 (O) due to merged cells or empty cols
            for col_idx in [13, 14, 15]:
                val = sheet.cell(row=row, column=col_idx).value
                if val is not None:
                    last_row = max(last_row, row)
                    if isinstance(val, str) and val.strip() not in ['Benchmarks', 'None', '', 'Lifetime', '1', '2', '3', '4', '5', '45', '> 1ms & <= 10ms', '> 10ms & <= 100ms', '> 100ms', '<= 1ms']:
                        bench_name = val.strip().lower().replace('_', '')
                        
                        # Find matching benchmark in df
                        matched_bench = None
                        for b in df['Benchmark'].unique():
                            norm_b = b.lower().replace('_', '')
                            if bench_name in norm_b or norm_b in bench_name:
                                matched_bench = b
                                break
                                
                        if matched_bench:
                            if matched_bench not in found_benches:
                                found_benches.append(matched_bench)
                            
                            # Write counts
                            for z in range(4):
                                for bkt in range(5):
                                    # Filter dataframe
                                    mask = (df['Benchmark'] == matched_bench) & (df['Zone'] == z) & (df['Bucket'] == bkt)
                                    count_series = df.loc[mask, 'Count']
                                    if not count_series.empty:
                                        count = count_series.values[0]
                                        sheet.cell(row=row+1+z, column=16+bkt).value = count
                                    else:
                                        sheet.cell(row=row+1+z, column=16+bkt).value = 0
                            break # Break inner column loop if matched
        
        # Now append any missing benchmarks below the last row
        current_row = last_row + 2
        
        from openpyxl.styles import Border, Side, Alignment, PatternFill
        thin_side = Side(style='thin')
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        blue_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        
        for b in df['Benchmark'].unique():
            if b not in found_benches:
                # Merge cells for benchmark name (span 4 rows)
                sheet.merge_cells(start_row=current_row+1, start_column=14, end_row=current_row+4, end_column=14)
                cell = sheet.cell(row=current_row+1, column=14)
                cell.value = b
                cell.alignment = center_align
                
                # Apply borders to the merged benchmark name block
                for r in range(current_row+1, current_row+5):
                    b_top = thin_side if r == current_row+1 else None
                    b_bottom = thin_side if r == current_row+4 else None
                    sheet.cell(row=r, column=14).border = Border(left=thin_side, right=thin_side, top=b_top, bottom=b_bottom)
                
                # Write row labels
                labels = ["<= 1ms", "> 1ms & <= 10ms", "> 10ms & <= 100ms", "> 100ms"]
                for i, label in enumerate(labels):
                    cell = sheet.cell(row=current_row+1+i, column=15)
                    # Handle if cell is merged from a previous incomplete run
                    if type(cell).__name__ == 'MergedCell':
                        for merged_range in sheet.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                sheet.unmerge_cells(str(merged_range))
                                cell = sheet.cell(row=current_row+1+i, column=15)
                                break
                    cell.value = label
                    cell.border = thin_border
                    cell.alignment = left_align
                    if i == 0:
                        cell.fill = blue_fill
                
                # Write counts
                for z in range(4):
                    for bkt in range(5):
                        mask = (df['Benchmark'] == b) & (df['Zone'] == z) & (df['Bucket'] == bkt)
                        count_series = df.loc[mask, 'Count']
                        
                        cell = sheet.cell(row=current_row+1+z, column=16+bkt)
                        if not count_series.empty:
                            cell.value = count_series.values[0]
                        else:
                            cell.value = 0
                            
                        cell.border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
                        cell.alignment = right_align
                        if z == 0:
                            cell.fill = blue_fill
                
                current_row += 5 # Skip 4 data rows + 1 empty spacing row
                            
        print(f"Total matched benchmarks in {sheet_name}: {len(found_benches)}")
    wb.save(excel_path)
    print(f"Successfully updated {excel_path}")

if __name__ == "__main__":
    filepath = 'collectStats/all_res'
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
        
    df = parse_all_res(filepath)
    if not df.empty:
        update_excel('Analysis_Motivation.xlsx', df)
    else:
        print("No MOT_STATS found.")
